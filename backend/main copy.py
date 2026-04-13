"""
BigBasket + Flipkart Reviews Scraper — FastAPI Backend
Patched v4 — Fixed BB page navigation stuck issue:
  - Reduced sleep times
  - Better stale detection
  - Faster page cycling
  - API-first approach with quick fallback
"""

import re, io, asyncio, random, json, sys, concurrent.futures
from datetime import datetime, timedelta, timezone
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page
from pydantic import BaseModel

def _run_in_proactor(coro):
    def _thread():
        if sys.platform == "win32":
            loop = asyncio.ProactorEventLoop()
            asyncio.set_event_loop(loop)
        else:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()
    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
        return pool.submit(_thread).result()

HEADLESS     = True
MAX_BB_PAGES = 150
MAX_FK_PAGES = 300

app = FastAPI(title="Review Scraper API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"],
                   allow_headers=["*"], allow_credentials=True, expose_headers=["Content-Disposition"])

class ScrapeRequest(BaseModel):
    url: str
    until_date: str = ""

_STEALTH = """
() => {
    Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
    if(!window.chrome)window.chrome={};
    window.chrome.runtime=window.chrome.runtime||{};window.chrome.app={isInstalled:false};
    window.chrome.csi=function(){};window.chrome.loadTimes=function(){};
    Object.defineProperty(navigator,'plugins',{get:()=>[
        {name:'Chrome PDF Plugin',filename:'internal-pdf-viewer'},
        {name:'Chrome PDF Viewer',filename:'mhjfbmdgcfjbbpaeojofohoefgiehjai'},
        {name:'Native Client',filename:'internal-nacl-plugin'},
        {name:'Widevine Content Decryption Module',filename:'widevinecdmadapter.dll'},
    ]});
    Object.defineProperty(navigator,'languages',{get:()=>['en-IN','en-US','en']});
    Object.defineProperty(navigator,'platform',{get:()=>'Win32'});
    Object.defineProperty(navigator,'vendor',{get:()=>'Google Inc.'});
    Object.defineProperty(screen,'width',{get:()=>1366});Object.defineProperty(screen,'height',{get:()=>768});
    window.devicePixelRatio=1;
    const gp=WebGLRenderingContext.prototype.getParameter;
    WebGLRenderingContext.prototype.getParameter=function(p){
        if(p===37445)return'Intel Inc.';if(p===37446)return'Intel Iris OpenGL Engine';return gp.call(this,p);
    };
    delete window.__playwright;delete window.__pw_manual;delete window.playwright;
}
"""
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.155 Safari/537.36"
_EH = {"Accept-Language":"en-IN,en-US;q=0.9,en;q=0.8","Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8","sec-ch-ua":'"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',"sec-ch-ua-mobile":"?0","sec-ch-ua-platform":'"Windows"'}

async def _make_ctx(pw):
    browser = await pw.chromium.launch(headless=HEADLESS,args=["--disable-blink-features=AutomationControlled","--no-sandbox","--disable-dev-shm-usage","--disable-web-security","--disable-features=VizDisplayCompositor,IsolateOrigins,site-per-process","--ignore-certificate-errors","--window-size=1366,768"])
    ctx = await browser.new_context(user_agent=_UA,locale="en-IN",timezone_id="Asia/Kolkata",viewport={"width":1366,"height":768},extra_http_headers=_EH)
    await ctx.add_init_script(_STEALTH)
    return browser, ctx

async def _scroll(page:Page,steps:int=4):
    for i in range(1,steps+1):
        await page.evaluate(f"window.scrollTo(0,document.body.scrollHeight*{i/steps})")
        await asyncio.sleep(0.2+random.uniform(0,0.1))

async def _goto(page:Page,url:str,retries:int=3)->bool:
    for attempt in range(retries):
        try:
            await page.goto(url,wait_until="domcontentloaded",timeout=60_000)
            try: await page.wait_for_load_state("networkidle",timeout=8_000)
            except: pass
            ln=await page.evaluate("()=>document.documentElement.outerHTML.length")
            if ln>2000: return True
        except Exception as e: print(f"[nav] attempt {attempt+1}: {e}")
        await asyncio.sleep(2*(attempt+1)+random.uniform(0,1))
    return False

_BF="""async([url,extra])=>{try{const r=await fetch(url,{method:'GET',credentials:'include',headers:{'Accept':'application/json,*/*',...extra}});if(!r.ok)return null;if(!(r.headers.get('content-type')||'').includes('json'))return null;return JSON.stringify(await r.json());}catch(e){return null;}}"""

async def _bfetch(page:Page,url:str,extra:dict=None)->Any:
    try:
        raw=await page.evaluate(_BF,[url,extra or {}])
        return json.loads(raw) if raw else None
    except: return None

COLUMNS=["review_id","rating","review_header","review_text","review_combine","review_country","date","product_name","brand","variant_name","sku","product_url","category","platform"]
COL_WIDTHS=[26,8,32,72,82,14,14,22,16,16,18,42,16,12]
COL_DESCS=[("review_id","Unique ID: Platform+Brand+Product+Date+Seq"),("rating","Star rating (1-5)"),("review_header","Title / headline of the review"),("review_text","Full body text of the review"),("review_combine","review_header + review_text concatenated"),("review_country","Country of review entry (default: india)"),("date","Review date (YYYY-MM-DD)"),("product_name","Full product name as listed on the platform"),("brand","Brand name of the product"),("variant_name","Pack size / variant (e.g. 10 kg, 500 ml)"),("sku","Platform SKU / Product ID"),("product_url","Direct URL of the product page"),("category","Product category (e.g. Atta, Oil, Rice)"),("platform","Source platform: bigbasket or flipkart")]
_SC={1:"C00000",2:"FF0000",3:"FFC000",4:"70AD47",5:"375623"}
_CKW=["Atta","Rice","Dal","Oil","Sugar","Salt","Flour","Spice","Masala","Biscuit","Snack","Milk","Ghee","Tea","Coffee","Soap","Shampoo","Detergent","Chips","Noodles","Pasta","Sauce","Pickle","Jam","Honey","Butter","Cheese","Curd","Yogurt","Juice","Drink","Water","Chocolate","Candy","Namkeen","Poha","Suji","Besan","Maida","Cornflour","Oats","Muesli","Cereals","Pulses"]

def _av(n:str)->str:
    m=re.search(r"\b(\d+(?:\.\d+)?\s*(?:kg|g|gm|ml|l|ltr|litre|oz|lb|pack|pc|pcs|piece|pieces|count|ct))\b",n or "",re.I)
    return m.group(0).strip() if m else ""

def _ac(url:str,n:str)->str:
    t=f"{url or ''} {n or ''}".lower()
    for kw in _CKW:
        if kw.lower() in t: return kw
    return ""

def make_review_id(prefix,brand,product,seq,date_part=""):
    p=prefix[:2].upper();br=(brand[:2] if brand else "XX").upper();pr=(product[:2] if product else "XX").upper()
    if not date_part: date_part=datetime.now().strftime("%d%m%Y")
    return f"{p}{br}{pr}{date_part}{seq:04d}"

def parse_until_date(s):
    if not s: return None
    for fmt in ("%d-%m-%Y","%Y-%m-%d"):
        try: return datetime.strptime(s.strip(),fmt).date()
        except: pass
    return None

def build_sheet_header(ws,color):
    hf=PatternFill("solid",start_color=color);hfn=Font(name="Calibri",bold=True,color="FFFFFF",size=11);ha=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for col,h in enumerate(COLUMNS,1):
        c=ws.cell(row=1,column=col,value=h);c.font,c.fill,c.alignment=hfn,hf,ha
    ws.row_dimensions[1].height=30

def write_rows(ws,rows_data,rating_col_idx=2):
    fe=PatternFill("solid",start_color="EBF3FB");fo=PatternFill("solid",start_color="FFFFFF");df=Font(name="Calibri",size=10);wrap=Alignment(wrap_text=True,vertical="top")
    for i,vals in enumerate(rows_data,1):
        fill=fe if i%2==0 else fo
        for col,val in enumerate(vals,1):
            cell=ws.cell(row=i+1,column=col,value=val);cell.font,cell.fill,cell.alignment=df,fill,wrap
            if col==rating_col_idx and isinstance(val,int) and val in _SC:
                cell.font=Font(name="Calibri",size=10,bold=True,color=_SC[val])

def apply_col_widths(ws):
    for col,w in enumerate(COL_WIDTHS,1): ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions

def add_column_reference_sheet(wb):
    ws=wb.create_sheet("Column Reference");hf=PatternFill("solid",start_color="2E4057");hfn=Font(name="Calibri",bold=True,color="FFFFFF",size=11);ha=Alignment(horizontal="center",vertical="center")
    for ci,label in enumerate(["Column Name","Description"],1):
        c=ws.cell(row=1,column=ci,value=label);c.font,c.fill,c.alignment=hfn,hf,ha
    ws.row_dimensions[1].height=25;alt1=PatternFill("solid",start_color="EBF3FB");alt2=PatternFill("solid",start_color="FFFFFF");wrap=Alignment(wrap_text=True,vertical="top")
    for i,(col,desc) in enumerate(COL_DESCS,2):
        fill=alt1 if i%2==0 else alt2
        c1=ws.cell(row=i,column=1,value=col);c1.font=Font(name="Calibri",size=10,bold=True);c1.fill,c1.alignment=fill,wrap
        c2=ws.cell(row=i,column=2,value=desc);c2.font=Font(name="Calibri",size=10);c2.fill,c2.alignment=fill,wrap
    ws.column_dimensions["A"].width=20;ws.column_dimensions["B"].width=65;ws.freeze_panes="A2"

def _xlsx(wb):
    buf=io.BytesIO();wb.save(buf);buf.seek(0);return buf.read()

# ══════════════════════════ BIGBASKET ══════════════════════════════════════════

def parse_bigbasket_url(url):
    for pat in [r"/product-reviews/(\d+)/([^/?#]+)",r"/pd/(\d+)/([^/?#]+)"]:
        m=re.search(pat,url)
        if m: return int(m.group(1)),m.group(2).strip("/")
    m=re.search(r"/pd/([^/\d][^/]*?)/(\d+)",url)
    if m: return int(m.group(2)),m.group(1).strip("/")
    m=re.search(r"-(\d{6,10})/?$",url.rstrip("/"))
    if m:
        sm=re.search(r"/product/([^/]+)",url);slug=sm.group(1).rsplit("-",1)[0] if sm else "product"
        return int(m.group(1)),slug
    raise ValueError("Could not extract product SKU from BigBasket URL.")

_bbs=[0]

def _bbmine(body,depth=0):
    if depth>7: return []
    out=[]
    if isinstance(body,list):
        for item in body:
            if isinstance(item,dict) and("review_id" in item or "rating" in item or "review_title" in item): out.append(item)
            elif isinstance(item,(dict,list)): out.extend(_bbmine(item,depth+1))
    elif isinstance(body,dict):
        for k in ("reviews","top_reviews","review_list","reviewList","data","results","items","content"):
            if isinstance(body.get(k),list): out.extend(_bbmine(body[k],depth+1))
        for k in ("rnr","review_data","reviewData","ratingReview","rating_review"):
            if body.get(k): out.extend(_bbmine(body[k],depth+1))
        for v in body.values():
            if isinstance(v,(dict,list)): out.extend(_bbmine(v,depth+1))
    return out

def _bbmine_str(raw):
    try: return _bbmine(json.loads(raw))
    except: return []

def _bbnorm(r,sku):
    rid=str(r.get("review_id") or r.get("id") or r.get("reviewId") or "")
    if not rid: _bbs[0]+=1;rid=f"bb_{_bbs[0]}"
    try: rating=float(str(r.get("rating") or r.get("star_rating") or r.get("stars") or ""))
    except: rating=""
    header=str(r.get("review_title") or r.get("title") or r.get("headline") or r.get("reviewTitle") or "").strip()
    text=str(r.get("review_description") or r.get("description") or r.get("body") or r.get("review_text") or r.get("reviewBody") or "").strip()
    if not header and not text and rating=="": return None
    raw_date=r.get("submitted_on") or r.get("published_on") or r.get("date") or r.get("created_at") or r.get("reviewDate") or ""
    try: date_val=datetime.fromisoformat(str(raw_date).replace("Z","+00:00")).strftime("%Y-%m-%d")
    except: date_val=str(raw_date)[:10] if raw_date else ""
    return {"review_id":rid,"rating":rating,"review_header":header,"review_text":text,"date":date_val,"product_name":str(r.get("product_name") or r.get("productName") or "").strip(),"brand":str(r.get("brand") or r.get("brandName") or "").strip(),"sku":str(r.get("sku") or r.get("product_id") or sku or ""),"category":str(r.get("category") or r.get("categoryName") or "").strip(),"variant_me":str(r.get("pack_desc") or r.get("size") or r.get("weight") or r.get("variant") or "").strip()}

def _bbadd(all_reviews,candidates,sku):
    for r in candidates:
        if not isinstance(r,dict): continue
        n=_bbnorm(r,sku)
        if n: all_reviews[n["review_id"]]=n

def _bbcutoff(reviews,until_date):
    if not until_date: return False
    for r in reviews.values():
        try:
            if datetime.fromisoformat(str(r.get("date","")).replace("Z","+00:00")).date()<until_date: return True
        except: pass
    return False

def _bbfilter(reviews,until_date):
    if not until_date: return reviews
    out={}
    for rid,r in reviews.items():
        try: keep=datetime.fromisoformat(str(r.get("date","")).replace("Z","+00:00")).date()>=until_date
        except: keep=True
        if keep: out[rid]=r
    return out

async def _bb_summary(page):
    return await page.evaluate(r"""
        ()=>{
            const s={};
            const re=document.querySelector('._1q4Li');if(re)s.overall_rating=re.textContent.trim().split('\n')[0].trim();
            const ce=document.querySelector('.gmwyk');
            if(ce){const t=ce.textContent.trim();const rm=t.match(/([\d,]+)\s*Ratings/);const rv=t.match(/([\d,]+)\s*Reviews/);s.total_ratings=rm?rm[1].replace(/,/g,''):'';s.total_reviews=rv?rv[1].replace(/,/g,''):'';}
            document.querySelectorAll('._2WvKG').forEach(el=>{const star=el.querySelector('._3WAZd'),cnt=el.querySelector('._3aK0L');if(star&&cnt)s['star_'+star.textContent.trim().replace(/\D/g,'')]=cnt.textContent.trim();});
            s.highlights=[];
            document.querySelectorAll('._28196').forEach(el=>{const score=el.querySelector('.xhZMN'),label=el.querySelector('._3v1xg'),ratings=el.querySelector('._3mmpM');if(score&&label)s.highlights.push({attribute:label.textContent.trim(),score:score.textContent.trim(),ratings:ratings?ratings.textContent.trim().replace(/\D/g,''):''});});
            const be=document.querySelector('._3TIT6');if(be)s.brand=be.textContent.trim();
            const ne=document.querySelector('.FkplV');if(ne)s.product_name=ne.textContent.trim();
            return s;
        }
    """)

async def scrape_bigbasket(sku, slug, until_date=None):
    _bbs[0] = 0
    all_reviews = {}
    intercepted = []
    summary = {}
    product_url = f"https://www.bigbasket.com/pd/{sku}/{slug}/"
    total_count = 0

    async def on_resp(resp):
        url = resp.url
        if "bigbasket.com" not in url or not re.search(r"(review|rnr|rating|feedback)", url, re.I): return
        try:
            ct = resp.headers.get("content-type", "")
            if "json" not in ct: return
            body = await resp.json()
            print(f"[BB] XHR: {url[:90]}")
            intercepted.append(body)
        except: pass

    async with async_playwright() as pw:
        browser, ctx = await _make_ctx(pw)
        page = await ctx.new_page()
        page.on("response", on_resp)

        # Warmup
        print("[BB] Warmup…")
        await _goto(page, "https://www.bigbasket.com/")
        await asyncio.sleep(random.uniform(1, 2))
        await _scroll(page, steps=2)
        await _goto(page, product_url)

        ms = re.search(r"/pd/\d+/([^/?#]+)", page.url)
        if ms:
            d = ms.group(1).strip("/")
            if d and d != slug:
                print(f"[BB] Slug: {slug!r}→{d!r}")
                slug = d
                product_url = f"https://www.bigbasket.com/pd/{sku}/{slug}/"

        await _scroll(page, steps=3)
        await asyncio.sleep(1)

        for body in list(intercepted): _bbadd(all_reviews, _bbmine(body), sku)
        intercepted.clear()
        print(f"[BB] After warmup XHR: {len(all_reviews)}")

        try:
            summary = await _bb_summary(page)
            summary["product_url"] = product_url
            tc = summary.get("total_reviews", "0")
            total_count = int(str(tc).replace(",", "")) if tc else 0
            print(f"[BB] Site total: {total_count}")
        except: summary = {"product_url": product_url}

        # ── Pure API approach (fastest) ──
        API = "https://www.bigbasket.com/catalog_svc/rnr/api/external/v1/review"
        HDRS = {"x-channel": "BB-WEB", "Referer": "https://www.bigbasket.com/", "Origin": "https://www.bigbasket.com"}

        # Top reviews first
        for n in (300, 100, 50):
            data = await _bfetch(page, f"{API}/top_reviews/{sku}?count={n}", HDRS)
            if data:
                hits = _bbmine(data)
                if hits:
                    before = len(all_reviews)
                    _bbadd(all_reviews, hits, sku)
                    print(f"[BB] top_reviews/{n}: +{len(all_reviews)-before}")
                    break
            await asyncio.sleep(0.3)

        # Page-by-page API — fast loop, no sleep between pages
        api_page = 1
        empty_runs = 0
        api_total = 0
        consecutive_empty = 0

        while True:
            urls = [
                f"{API}/get_reviews?product_id={sku}&page={api_page}&count=20&sort_by=recent",
                f"{API}/get_reviews?product_id={sku}&page={api_page}&page_size=20",
                f"{API}/get_reviews?product_id={sku}&offset={(api_page-1)*20}&limit=20",
            ]
            got = False
            for u in urls:
                data = await _bfetch(page, u, HDRS)
                if not data: continue
                if not api_total:
                    api_total = data.get("total_count") or data.get("total") or 0
                hits = _bbmine(data)
                if not hits: continue
                before = len(all_reviews)
                _bbadd(all_reviews, hits, sku)
                added = len(all_reviews) - before
                if added > 0:
                    print(f"[BB] API p{api_page}: +{added} | total={len(all_reviews)}")
                    got = True
                    consecutive_empty = 0
                    break
                await asyncio.sleep(0.1)

            if not got:
                consecutive_empty += 1
                print(f"[BB] API p{api_page}: empty ({consecutive_empty})")
                if consecutive_empty >= 5:
                    print("[BB] 5 consecutive empty — API done.")
                    break
            else:
                if _bbcutoff(all_reviews, until_date):
                    print("[BB] Cutoff reached.")
                    break
                if api_total and len(all_reviews) >= api_total:
                    print(f"[BB] Got all {api_total} via API!")
                    break

            api_page += 1
            # Very short sleep — just enough to avoid rate limit
            await asyncio.sleep(random.uniform(0.3, 0.8))

        print(f"[BB] After API: {len(all_reviews)} reviews")

        # If API got enough, done!
        if (api_total > 0 and len(all_reviews) >= api_total * 0.9) or \
           (not api_total and len(all_reviews) >= 30 and consecutive_empty >= 5):
            print("[BB] API sufficient — done!")
            await browser.close()
            filtered = _bbfilter(all_reviews, until_date)
            return summary, list(filtered.values()), api_total or len(filtered)

        # ── Phase 2: Page navigation fallback ──
        print("[BB] Phase 2: page nav fallback…")
        review_base = f"https://www.bigbasket.com/product-reviews/{sku}/{slug}/"
        stale = 0

        for pg in range(1, MAX_BB_PAGES + 1):
            print(f"[BB] → Page {pg}")
            ok = await _goto(page, f"{review_base}?page={pg}")
            if not ok:
                stale += 1
                if stale >= 3: break
                continue

            await _scroll(page, steps=3)
            prev = len(all_reviews)

            # Grab XHR intercepts
            for body in list(intercepted): _bbadd(all_reviews, _bbmine(body), sku)
            intercepted.clear()

            added = len(all_reviews) - prev
            print(f"[BB] Page {pg}: +{added} | total={len(all_reviews)}")

            stale = 0 if added > 0 else stale + 1
            if stale >= 3:
                print("[BB] 3 stale pages — stop.")
                break
            if _bbcutoff(all_reviews, until_date):
                print("[BB] Cutoff.")
                break
            if total_count and len(all_reviews) >= total_count:
                print(f"[BB] Got all {total_count}.")
                break

            await asyncio.sleep(random.uniform(1.5, 3.0))

        await browser.close()

    filtered = _bbfilter(all_reviews, until_date)
    return summary, list(filtered.values()), total_count or len(filtered)

def build_bigbasket_rows(reviews, slug, product_url="", summary=None):
    summary = summary or {}
    fb_name = summary.get("product_name", "")
    fb_brand = summary.get("brand", "")
    dp = datetime.now().strftime("%d%m%Y")
    sm = re.search(r"/pd/(\d+)/", product_url or "") or re.search(r"/product-reviews/(\d+)/", product_url or "")
    sku_url = sm.group(1) if sm else ""
    slug_name = slug.replace("-", " ").title() if slug else ""
    slug_var = _av(slug.replace("-", " ")) if slug else ""
    rows = []
    for i, r in enumerate(reviews, 1):
        name = r.get("product_name") or fb_name or slug_name or ""
        brand = r.get("brand") or fb_brand or ""
        if not brand and slug:
            brand = slug.split("-")[0].title()
        rh = r.get("review_header") or ""
        rt = r.get("review_text") or ""
        var = r.get("variant_me") or r.get("pack_desc") or _av(name) or slug_var or ""
        raw_s = str(r.get("sku") or "").strip()
        sku = raw_s if raw_s and raw_s != "0" else sku_url
        raw_c = (r.get("category") or "").strip()
        cat = raw_c or _ac(product_url, name) or _ac(slug, slug_name)
        rows.append([make_review_id("BB", brand, slug, i, dp), r.get("rating", ""), rh, rt,
                     f"{rh} {rt}".strip(), "india", r.get("date", ""), name, brand, var, sku,
                     product_url, cat, "bigbasket"])
    return rows

def build_bigbasket_excel(reviews, summary, slug, product_url=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "BigBasket Reviews"
    build_sheet_header(ws, "1F4E79")
    write_rows(ws, build_bigbasket_rows(reviews, slug, product_url, summary))
    apply_col_widths(ws)
    ws2 = wb.create_sheet("Summary")
    rows2 = [
        ("Product URL", summary.get("product_url", product_url)),
        ("Overall Rating", summary.get("overall_rating", "")),
        ("Total Ratings", summary.get("total_ratings", "")),
        ("Total Reviews", summary.get("total_reviews", "")),
        ("5 Star", summary.get("star_5", "")),
        ("4 Star", summary.get("star_4", "")),
        ("3 Star", summary.get("star_3", "")),
        ("2 Star", summary.get("star_2", "")),
        ("1 Star", summary.get("star_1", "")),
        ("Reviews Scraped", len(reviews)),
        ("Scrape Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for h in summary.get("highlights", []):
        rows2 += [(f"{h['attribute']} Score", h.get("score", "")),
                  (f"{h['attribute']} Ratings", h.get("ratings", ""))]
    hf = Font(name="Calibri", bold=True, size=10)
    df = Font(name="Calibri", size=10)
    for i, (f, v) in enumerate(rows2, 1):
        ws2.cell(row=i, column=1, value=f).font = hf
        ws2.cell(row=i, column=2, value=v).font = df
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40
    add_column_reference_sheet(wb)
    return _xlsx(wb)

# ══════════════════════════ FLIPKART ══════════════════════════════════════════

def parse_relative_date(raw):
    today = datetime.now(tz=timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    raw = raw.strip().lstrip("·").strip().lower()
    m = re.match(r"(\d+)\s+(day|week|month|year)s?\s+ago", raw)
    if m:
        n, unit = int(m.group(1)), m.group(2)
        return today - {"day": timedelta(days=n), "week": timedelta(weeks=n),
                        "month": timedelta(days=n*30), "year": timedelta(days=n*365)}[unit]
    for fmt in ("%d %b %Y", "%d %B %Y", "%b %d, %Y", "%b, %Y", "%b %Y"):
        try: return datetime.strptime(raw, fmt)
        except: pass
    return None

def parse_flipkart_url(url):
    m = re.search(r"flipkart\.com/([^/]+)/product-reviews/([^?/]+)", url)
    if m:
        pid = re.search(r"pid=([A-Z0-9]+)", url)
        lid = re.search(r"lid=([A-Z0-9]+)", url)
        return m.group(1), m.group(2), pid.group(1) if pid else None, lid.group(1) if lid else None
    m2 = re.search(r"flipkart\.com/([^/]+)/p/([^?/]+)", url)
    if m2:
        pid = re.search(r"pid=([A-Z0-9]+)", url)
        lid = re.search(r"lid=([A-Z0-9]+)", url)
        return m2.group(1), m2.group(2), pid.group(1) if pid else None, lid.group(1) if lid else None
    raise ValueError("Could not parse Flipkart URL.")

DUMP_JS = r"""
()=>{
    const cards=[],seen=new Set();
    function getRating(el){for(const a of el.querySelectorAll('[aria-label]')){const lbl=(a.getAttribute('aria-label')||'').toLowerCase();const m=lbl.match(/(\d(?:\.\d)?)\s*(?:out of|\/)\s*5/)||lbl.match(/rated\s*(\d(?:\.\d)?)/)||lbl.match(/^(\d(?:\.\d)?)\s*star/);if(m){const v=parseFloat(m[1]);if(v>=1&&v<=5)return Math.round(v);}}for(const a of el.querySelectorAll('[data-score],[data-rating],[data-value]')){const v=parseFloat(a.getAttribute('data-score')||a.getAttribute('data-rating')||a.getAttribute('data-value'));if(!isNaN(v)&&v>=1&&v<=5)return Math.round(v);}for(const a of el.querySelectorAll('span,div')){if(a.children.length>0)continue;const t=(a.textContent||'').trim();if(!/^[1-5](\.[05])?$/.test(t))continue;const bg=window.getComputedStyle(a).backgroundColor||'';if(bg&&bg!=='rgba(0, 0, 0, 0)'&&bg!=='rgb(255, 255, 255)'){const v=parseFloat(t);if(v>=1&&v<=5)return Math.round(v);}}for(const a of el.querySelectorAll('span,div,b,strong')){if(a.children.length>0)continue;const t=(a.textContent||'').trim();if(/^[1-5]$/.test(t)&&a.offsetWidth<60&&a.offsetHeight<60)return parseInt(t);}return 0;}
    for(const card of document.querySelectorAll('div.css-g5y9jx')){const s=card.getAttribute('style')||'';if(!s.includes('padding-left: 16px')||!s.includes('padding-top: 16px'))continue;const t=(card.innerText||'').trim();if(t.length<40)continue;const k=t.slice(0,150);if(!seen.has(k)){seen.add(k);cards.push({rating:getRating(card),text:t});}}
    if(cards.length>=3)return JSON.stringify(cards);
    cards.length=0;seen.clear();
    for(const body of document.querySelectorAll('div.css-146c3p1')){const bt=(body.innerText||'').trim();if(bt.length<30)continue;let cur=body.parentElement;for(let i=0;i<10&&cur&&cur!==document.body;i++){const t=(cur.innerText||'').trim();if(t.length>100&&t.length<3000){const k=t.slice(0,150);if(!seen.has(k)){seen.add(k);cards.push({rating:getRating(cur),text:t});}break;}cur=cur.parentElement;}}
    if(cards.length>=3)return JSON.stringify(cards);
    const MONTH=/\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b/i;cards.length=0;seen.clear();
    for(const el of document.querySelectorAll('div')){const t=(el.innerText||'').trim();if(t.length<80||t.length>3000||!MONTH.test(t))continue;const child=[...el.querySelectorAll('div')].some(d=>{const dt=(d.innerText||'').trim();return dt.length>=80&&dt.length<=3000&&MONTH.test(dt);});if(child)continue;const k=t.slice(0,150);if(!seen.has(k)){seen.add(k);cards.push({rating:getRating(el),text:t});}}
    return JSON.stringify(cards);
}
"""

_FKP = r"""
()=>{
    const info={product_name:'',brand:'',variant_me:'',sku:'',category:''};
    const h1=document.querySelector('h1.yhB1nd,h1[class*="title"],h1[class*="product"],._9E25nV span,.B_NuCI');if(h1)info.product_name=h1.textContent.trim();
    if(!info.product_name){const p=document.title.split(/[|\-–]/);if(p.length>0)info.product_name=p[0].trim();}
    const be=document.querySelector('._2whKao,[class*="brand"],.G6XhRU');if(be)info.brand=be.textContent.trim();
    for(const s of document.querySelectorAll('script[type="application/ld+json"]')){try{const data=JSON.parse(s.textContent||'{}');for(const item of(Array.isArray(data)?data:[data])){if(!info.brand&&item.brand)info.brand=typeof item.brand==='string'?item.brand:(item.brand.name||'');if(!info.sku&&item.sku)info.sku=item.sku;if(!info.product_name&&item.name)info.product_name=item.name;if(!info.category&&item.category)info.category=item.category;if(!info.category&&item['@type']==='BreadcrumbList'){const cats=(item.itemListElement||[]).map(x=>x.name||'').filter(Boolean);if(cats.length>1)info.category=cats.slice(1,-1).join(' > ');}}}catch(e){}}
    const sv=document.querySelector('._3Rm2xQ._1MR4o5,._1v_viwn._2BXCoD');if(sv)info.variant_me=sv.textContent.trim();
    if(!info.variant_me){const sm=(info.product_name||'').match(/\b(\d+\s*(?:g|kg|ml|l|ltr|litre|gm|oz|lb|pack|pc|pcs|piece|pieces|count|ct))\b/i);if(sm)info.variant_me=sm[0].trim();}
    if(!info.category){const bcs=[...document.querySelectorAll('._2whKao a,._3GIHBu a,[class*="breadcrumb"] a,nav a')];const cats=bcs.map(a=>a.textContent.trim()).filter(t=>t&&t!=='Home'&&t.length<50);if(cats.length>0)info.category=cats.join(' > ');}
    if(!info.sku){const pid=(window.location.search||'').match(/pid=([A-Z0-9]+)/);if(pid)info.sku=pid[1];}
    info.brand=info.brand.replace(/visit the store/gi,'').trim();return info;
}
"""

def parse_flipkart_card(rating, text, media=None):
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    month_re = re.compile(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b", re.I)
    JUNK = {"READ MORE","Certified Buyer","Like","Dislike","Report Abuse","HELPFUL","Was this review helpful","Thank you for your feedback"}
    def is_junk(l): return l in JUNK or bool(re.fullmatch(r"[\d,\.★\s]+", l)) or len(l) < 2
    date_str = next((l for l in lines if month_re.search(l) and len(l) < 35), "")
    parsed_date = ""
    if date_str:
        dt = parse_relative_date(date_str)
        parsed_date = dt.strftime("%Y-%m-%d") if dt else date_str.strip()
    if not rating:
        for l in lines:
            if re.match(r"^([1-5])(\.\d)?$", l.strip()):
                try: rating = int(l.strip()[0]); break
                except: pass
    rem = [l for l in lines if not is_junk(l) and not month_re.search(l)]
    title = next((l for l in rem if 4 <= len(l) <= 140), "")
    body_lines, seen_b = [], set()
    for l in lines:
        if is_junk(l) or month_re.search(l) or l == title: continue
        if l.startswith("Review for:") or "Verified Purchase" in l: continue
        if re.fullmatch(r"[1-5](\.[05])?", l.strip()): continue
        if len(l) > 8 and l not in seen_b: seen_b.add(l); body_lines.append(l)
    return {"rating": rating, "title": title, "body": " ".join(body_lines), "date_str": parsed_date, "reviewer": "", "media": media or []}

def _fkpu(url):
    m = re.search(r"(flipkart\.com/)([^/]+)/product-reviews/([^?]+)(.*)", url)
    if m:
        pid = re.search(r"pid=([A-Z0-9]+)", m.group(4))
        return f"https://www.{m.group(1)}{m.group(2)}/p/{m.group(3)}" + (f"?pid={pid.group(1)}" if pid else "")
    return ""

async def scrape_flipkart(url, until_date=None):
    slug, item, pid, lid = parse_flipkart_url(url)
    all_reviews = []
    seen_keys = set()
    product_info = {}

    def rev_url(pg):
        base = f"https://www.flipkart.com/{slug}/product-reviews/{item}?marketplace=FLIPKART&sortOrder=MOST_RECENT&page={pg}"
        if pid: base += f"&pid={pid}"
        if lid: base += f"&lid={lid}"
        return base

    prod_url = _fkpu(url) or (f"https://www.flipkart.com/{slug}/p/{item}" + (f"?pid={pid}" if pid else ""))

    async with async_playwright() as pw:
        browser, ctx = await _make_ctx(pw)
        page = await ctx.new_page()
        try:
            await _goto(page, "https://www.flipkart.com/")
            await asyncio.sleep(random.uniform(1, 2))
            for txt in ["✕", "×", "Close"]:
                try: await page.get_by_text(txt, exact=True).first.click(timeout=1500); break
                except: pass
            await page.keyboard.press("Escape")
        except Exception as e: print(f"[FK] Warmup: {e}")

        try:
            await _goto(page, prod_url)
            await asyncio.sleep(2)
            result = await page.evaluate(_FKP)
            if result: product_info = {k: v for k, v in result.items() if v}
            print(f"[FK] Product: {product_info.get('product_name', '')!r}")
        except Exception as e: print(f"[FK] Product info: {e}")

        ce = 0
        for pg in range(1, MAX_FK_PAGES + 1):
            ok = await _goto(page, rev_url(pg))
            if not ok: ce += 1
            if ce >= 5: break
            await asyncio.sleep(random.uniform(1.5, 2.5))
            await _scroll(page)
            try: raw = await page.evaluate(DUMP_JS); cards = json.loads(raw) if raw else []
            except: cards = []
            new = 0; rc = False
            for card in cards:
                r = parse_flipkart_card(card["rating"], card["text"], card.get("media", []))
                bk = (r["body"] or "")[:300].strip()
                tk = (r["title"] or "")[:100].strip()
                key = (tk, bk, r["date_str"], r["rating"])
                if not (bk or tk or r["rating"]): continue
                if key in seen_keys: continue
                if until_date and r["date_str"]:
                    try:
                        if datetime.strptime(r["date_str"], "%Y-%m-%d").date() < until_date: rc = True; continue
                    except: pass
                seen_keys.add(key); new += 1; all_reviews.append(r)
            print(f"[FK] Page {pg}: +{new} | total={len(all_reviews)}")
            if rc: break
            ce = 0 if new > 0 else ce + 1
            if ce >= 5: break
        await browser.close()
    return all_reviews, product_info

def build_flipkart_rows(reviews, product_info, product_url=""):
    name = product_info.get("product_name", "")
    name = re.sub(r"(?i)\s*price\s+in\s+india", "", name).strip()
    brand = product_info.get("brand", "")
    slug_from_url = re.search(r"flipkart\.com/([^/]+)/", product_url or "")
    slug_str = slug_from_url.group(1).replace("-", " ") if slug_from_url else ""
    var = product_info.get("variant_me", "").strip() or _av(name) or _av(slug_str)
    rs = str(product_info.get("sku", "") or "").strip()
    skm = re.search(r"/p/([^/?]+)", str(product_url))
    pid_m = re.search(r"pid=([A-Z0-9]+)", str(product_url))
    sku = (rs if rs and rs != "0" else "") or (pid_m.group(1) if pid_m else "") or (skm.group(1) if skm else "")
    cat = (product_info.get("category", "") or "").strip() or _ac(product_url, name)
    dp = datetime.now().strftime("%d%m%Y")
    date_list = [r.get("date_str", "") for r in reviews if r.get("date_str", "")]
    fallback_date = max(set(date_list), key=date_list.count) if date_list else ""
    rows = []
    for i, r in enumerate(reviews, 1):
        rh = r.get("title") or ""
        rt = r.get("body") or ""
        date_val = r.get("date_str", "") or fallback_date
        rows.append([make_review_id("FK", brand or "FL", name or "PR", i, dp), r["rating"], rh, rt,
                     f"{rh} {rt}".strip(), "india", date_val, name, brand, var, sku,
                     product_url, cat, "flipkart"])
    return rows

def build_flipkart_excel(reviews, product_info=None, product_url=""):
    product_info = product_info or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Flipkart Reviews"
    build_sheet_header(ws, "F47920")
    write_rows(ws, build_flipkart_rows(reviews, product_info, product_url))
    apply_col_widths(ws)
    ws2 = wb.create_sheet("Summary")
    df_r = [r["rating"] for r in reviews if r.get("rating")]
    avg = round(sum(df_r) / len(df_r), 2) if df_r else ""
    dates = sorted(r["date_str"] for r in reviews if r.get("date_str"))
    rows2 = [
        ("Product URL", product_url),
        ("Product Name", product_info.get("product_name", "")),
        ("Brand", product_info.get("brand", "")),
        ("Variant / Size", product_info.get("variant_me", "") or _av(product_info.get("product_name", ""))),
        ("SKU / PID", product_info.get("sku", "")),
        ("Category", product_info.get("category", "") or _ac(product_url, product_info.get("product_name", ""))),
        ("Scrape Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Reviews Scraped", len(reviews)),
        ("Average Rating (scraped)", avg),
        ("Date Range", f"{dates[0]} → {dates[-1]}" if dates else ""),
    ]
    hf = Font(name="Calibri", bold=True, size=10)
    df = Font(name="Calibri", size=10)
    for i, (f, v) in enumerate(rows2, 1):
        ws2.cell(row=i, column=1, value=f).font = hf
        ws2.cell(row=i, column=2, value=v).font = df
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50
    add_column_reference_sheet(wb)
    return _xlsx(wb)

# ══════════════════════════ ENDPOINTS ═════════════════════════════════════════

@app.get("/health")
async def health(): return {"status": "ok", "version": "4.0.0"}

@app.post("/api/validate-bigbasket")
@app.get("/api/validate-bigbasket")
async def validate_bigbasket(req: ScrapeRequest = None, url: str = ""):
    try:
        u = (req.url if req else None) or url or ""
        if not u: return {"valid": False, "error": "No URL provided"}
        sku, slug = parse_bigbasket_url(u)
        return {"valid": True, "sku": sku, "slug": slug}
    except Exception as e: return {"valid": False, "error": str(e)}

@app.post("/api/validate-flipkart")
@app.get("/api/validate-flipkart")
async def validate_flipkart(req: ScrapeRequest = None, url: str = ""):
    try:
        u = (req.url if req else None) or url or ""
        if not u: return {"valid": False, "error": "No URL provided"}
        slug, item, pid, lid = parse_flipkart_url(u)
        return {"valid": True, "slug": slug, "item": item}
    except Exception as e: return {"valid": False, "error": str(e)}

@app.post("/api/scrape-bigbasket")
async def scrape_bigbasket_endpoint(req: ScrapeRequest):
    try: sku, slug = parse_bigbasket_url(req.url)
    except ValueError as e: raise HTTPException(400, str(e))
    until = parse_until_date(req.until_date) if req.until_date else None
    try:
        summary, reviews, total = await asyncio.get_event_loop().run_in_executor(
            None, lambda: _run_in_proactor(scrape_bigbasket(sku, slug, until)))
    except Exception as e:
        import traceback
        raise HTTPException(500, f"Scraping failed: {type(e).__name__}: {e}\n{traceback.format_exc()}")
    if not reviews: raise HTTPException(404, "No reviews found.")
    data = build_bigbasket_excel(reviews, summary, slug, req.url)
    filename = f"BigBasket_{slug[:30]}_{len(reviews)}_reviews.xlsx"
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@app.post("/api/scrape-flipkart")
async def scrape_flipkart_endpoint(req: ScrapeRequest):
    try: parse_flipkart_url(req.url)
    except ValueError as e: raise HTTPException(400, str(e))
    until = parse_until_date(req.until_date) if req.until_date else None
    try:
        reviews, product_info = await asyncio.get_event_loop().run_in_executor(
            None, lambda: _run_in_proactor(scrape_flipkart(req.url, until)))
    except Exception as e:
        import traceback
        raise HTTPException(500, f"Scraping failed: {type(e).__name__}: {e}\n{traceback.format_exc()}")
    if not reviews: raise HTTPException(404, "No reviews found.")
    data = build_flipkart_excel(reviews, product_info, req.url)
    filename = f"Flipkart_{len(reviews)}_reviews.xlsx"
    return StreamingResponse(io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})