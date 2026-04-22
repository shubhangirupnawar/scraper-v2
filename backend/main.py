import asyncio
import sys

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from dotenv import load_dotenv
load_dotenv()

"""
BigBasket + Flipkart Reviews Scraper � FastAPI Backend
Merged best logic from:
  - bigbasket/scrapping.py  ? ReviewAPICapture, scroll loop, until_date, product summary
  - flipkart/scrap.py       ? parse_relative_date, HTML parsing, DUMP_JS card extractor
  - scraper-v2/main.py      ? FastAPI structure, Excel builder, unified column schema
"""

import os
from supabase import create_client, Client

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

import re, time, io, asyncio, random, json, sys
from datetime import datetime, timedelta
from urllib.parse import urljoin

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page, Request

app = FastAPI(title="Review Scraper API")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)


class ScrapeRequest(BaseModel):
    url: str
    until_date: str = ""   # optional DD-MM-YYYY or YYYY-MM-DD


# -- Shared column schema -------------------------------------------------------

COLUMNS = [
    "review_id", "rating", "review_header", "review_text", "review_combine",
    "review_entry", "date", "product_name", "brand",
    "variant_me", "sku", "product_url", "category", "platform"
]

COL_WIDTHS = [22, 8, 30, 70, 80, 14, 14, 20, 15, 14, 14, 40, 15, 12]

def save_reviews_to_supabase(rows: list):
    records = [dict(zip(COLUMNS, row)) for row in rows]
    try:
        supabase.table("reviews").insert(records).execute()
        print(f"[Supabase] Saved {len(records)} reviews")
    except Exception as e:
        print(f"[Supabase] Error: {e}")


def make_review_id(platform_prefix: str, brand: str, product: str, seq: int) -> str:
    plat = platform_prefix[:2].upper()
    br = (brand[:2] if brand else "XX").upper()
    pr = (product[:2] if product else "XX").upper()
    date_part = datetime.now().strftime("%d%m%Y")
    return f"{plat}{br}{pr}{date_part}{seq:04d}"


def build_sheet_header(ws, platform_color: str):
    hfill = PatternFill("solid", start_color=platform_color)
    hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, ha
    ws.row_dimensions[1].height = 30


def write_rows(ws, rows_data: list, rating_col_idx: int = 2):
    fe = PatternFill("solid", start_color="EBF3FB")
    fo = PatternFill("solid", start_color="FFFFFF")
    df = Font(name="Calibri", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    sc = {1: "C00000", 2: "FF0000", 3: "FFC000", 4: "70AD47", 5: "375623"}
    for i, vals in enumerate(rows_data, 1):
        row = i + 1
        fill = fe if i % 2 == 0 else fo
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font, cell.fill, cell.alignment = df, fill, wrap
            if col == rating_col_idx and isinstance(val, int) and val in sc:
                cell.font = Font(name="Calibri", size=10, bold=True, color=sc[val])


def apply_col_widths(ws):
    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def parse_until_date(date_str: str):
    """Parse date string in DD-MM-YYYY or YYYY-MM-DD, return date object or None."""
    if not date_str:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except ValueError:
            pass
    return None


# ------------------------------------------------------------------------------
# BIGBASKET
# ------------------------------------------------------------------------------

def parse_bigbasket_url(url: str):
    # /product-reviews/SKU/slug
    m = re.search(r"/product-reviews/(\d+)/([^/?]+)", url)
    if m:
        return int(m.group(1)), m.group(2)
    # /pd/SKU/slug
    m = re.search(r"/pd/(\d+)/([^/?]+)", url)
    if m:
        return int(m.group(1)), m.group(2).rstrip("/")
    # /pd/slug/SKU  (legacy)
    m = re.search(r"/pd/([^/\d][^/]*?)/(\d+)", url)
    if m:
        return int(m.group(2)), m.group(1)
    # trailing -SKU
    m = re.search(r"-(\d{6,10})/?$", url.rstrip("/"))
    if m:
        slug_m = re.search(r"/product/([^/]+)", url)
        slug = slug_m.group(1).rsplit("-", 1)[0] if slug_m else "product"
        return int(m.group(1)), slug
    raise ValueError("Could not extract product SKU from BigBasket URL.")


# -- Shared date helpers (from bigbasket/scrapping.py) -------------------------

def _bb_reached_cutoff(reviews_dict: dict, until_date) -> bool:
    if until_date is None:
        return False
    for r in reviews_dict.values():
        pub = r.get("date") or ""
        if not pub:
            continue
        try:
            d = datetime.fromisoformat(str(pub).replace("Z", "+00:00")).date()
            if d < until_date:
                return True
        except Exception:
            pass
    return False


def _bb_filter_by_date(reviews_dict: dict, until_date) -> dict:
    if until_date is None:
        return reviews_dict
    filtered = {}
    for rid, r in reviews_dict.items():
        pub = r.get("date") or ""
        keep = True
        if pub:
            try:
                d = datetime.fromisoformat(str(pub).replace("Z", "+00:00")).date()
                keep = d >= until_date
            except Exception:
                pass
        if keep:
            filtered[rid] = r
    return filtered


# -- Network interceptor (from bigbasket/scrapping.py) -------------------------

class ReviewAPICapture:
    def __init__(self):
        self.api_calls: list[dict] = []
        self.api_calls_with_body: list[dict] = []

    def on_request(self, request):
        url = request.url
        if re.search(r"(review|rnr)", url, re.IGNORECASE) and "bigbasket.com" in url:
            if url not in [c.get("url") for c in self.api_calls]:
                self.api_calls.append({"url": url, "headers": dict(request.headers)})

    async def on_response(self, response):
        url = response.url
        if re.search(r"(review|rnr)", url, re.IGNORECASE) and "bigbasket.com" in url:
            try:
                body = await response.json()
                self.api_calls_with_body.append({"url": url, "body": body})
            except Exception:
                pass


def _bb_normalise_review(r: dict) -> dict:
    raw_date = (r.get("submitted_on") or r.get("published_on") or "")
    try:
        date_val = datetime.fromisoformat(str(raw_date).replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except Exception:
        date_val = str(raw_date)[:10] if raw_date else ""
    return {
        "review_id":     r.get("review_id", ""),
        "rating":        r.get("rating", ""),
        "review_header": (r.get("review_title", "") or "").strip(),
        "review_text":   (r.get("review_description", "") or "").strip(),
        "date":          date_val,
        "product_name":  r.get("product_name", ""),
        "brand":         r.get("brand", ""),
        "sku":           r.get("sku", ""),
        "category":      r.get("category", ""),
    }


def _bb_parse_reviews_from_state(state_json: str) -> list[dict]:
    reviews = []
    try:
        data = json.loads(state_json)
        rnr = data.get("rnr", {})
        for section in ("top_reviews", "reviews"):
            for r in rnr.get(section, []):
                reviews.append(r)
    except Exception:
        pass
    return reviews


async def _bb_scrape_product_summary(page: Page) -> dict:
    """Extract overall rating, star breakdown, highlights from review page DOM."""
    return await page.evaluate("""
        () => {
            const s = {};
            const ratingEl = document.querySelector('._1q4Li');
            if (ratingEl) s.overall_rating = ratingEl.textContent.trim().split('\\n')[0].trim();
            const countEl = document.querySelector('.gmwyk');
            if (countEl) {
                const txt = countEl.textContent.trim();
                const rm = txt.match(/([\\d,]+)\\s*Ratings/);
                const rv = txt.match(/([\\d,]+)\\s*Reviews/);
                s.total_ratings = rm ? rm[1].replace(/,/g,'') : '';
                s.total_reviews = rv ? rv[1].replace(/,/g,'') : '';
            }
            document.querySelectorAll('._2WvKG').forEach(el => {
                const star = el.querySelector('._3WAZd');
                const count = el.querySelector('._3aK0L');
                if (star && count) {
                    const n = star.textContent.trim().replace(/\\D/g,'');
                    s['star_' + n] = count.textContent.trim();
                }
            });
            s.highlights = [];
            document.querySelectorAll('._28196').forEach(el => {
                const score = el.querySelector('.xhZMN');
                const label = el.querySelector('._3v1xg');
                const ratings = el.querySelector('._3mmpM');
                if (score && label) {
                    s.highlights.push({
                        attribute: label.textContent.trim(),
                        score: score.textContent.trim(),
                        ratings: ratings ? ratings.textContent.trim().replace(/\\D/g,'') : ''
                    });
                }
            });
            const brandEl = document.querySelector('._3TIT6');
            if (brandEl) s.brand = brandEl.textContent.trim();
            const nameEl = document.querySelector('.FkplV');
            if (nameEl) s.product_name = nameEl.textContent.trim();
            return s;
        }
    """)


async def scrape_bigbasket(sku: int, slug: str, until_date=None):
    """
    Scrape BigBasket reviews using:
    1. Network interception (ReviewAPICapture) � fastest
    2. window.__PRELOADED_STATE__ parsing
    3. Scroll loop for lazy-loaded reviews
    4. Paginated API fetch using captured endpoint
    5. DOM fallback if all else fails
    """
    all_reviews: dict[int, dict] = {}   # keyed by review_id for dedup
    product_url = f"https://www.bigbasket.com/pd/{sku}/{slug}/"
    review_url_base = f"https://www.bigbasket.com/product-reviews/{sku}/{slug}/"
    summary = {}

    capture = ReviewAPICapture()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-sandbox',
                '--disable-dev-shm-usage',
            ]
        )
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-IN",
            viewport={"width": 1280, "height": 900},
        )
        page: Page = await ctx.new_page()
        await ctx.add_init_script("""
            Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
            window.chrome={runtime:{}};
            Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});
            Object.defineProperty(navigator,'languages',{get:()=>['en-IN','en']});
        """)

        # Wire network listeners BEFORE navigation
        page.on("request", capture.on_request)
        page.on("response", capture.on_response)

        # Skip homepage, go directly to product page first to set cookies
        print("[BB] Loading product page to set cookies...")
        await page.goto(f"https://www.bigbasket.com/pd/{sku}/{slug}/", wait_until="domcontentloaded", timeout=60_000)
        await asyncio.sleep(4)

        # Now load review page
        start_url = f"{review_url_base}?page=1"
        print(f"[BB] Loading review page: {start_url}")
        await page.goto(start_url, wait_until="domcontentloaded", timeout=60_000)
        await asyncio.sleep(6)

        # Check if blocked
        page_content = await page.content()
        if "Something went wrong" in page_content or "Internal Server" in page_content:
            print("[BB] Blocked on review page, retrying after delay...")
            await asyncio.sleep(5)
            await page.reload(wait_until="domcontentloaded", timeout=60_000)
            await asyncio.sleep(6)

        # Scrape product summary from DOM
        try:
            summary = await _bb_scrape_product_summary(page)
            summary["product_url"] = product_url
            print(f"[BB] Summary: {summary.get('overall_rating')} rating, {summary.get('total_reviews')} reviews")
        except Exception as e:
            print(f"[BB] Summary error: {e}")
            summary = {"product_url": product_url}

        # Parse __PRELOADED_STATE__ for initial reviews
        state_str = await page.evaluate("""
            () => {
                try {
                    const s = window.__PRELOADED_STATE__;
                    if (!s) return '{}';
                    if (typeof s === 'string') return s;
                    return JSON.stringify(s);
                }
                catch(e) { return '{}'; }
            }
        """)
        total_count = 0
        try:
            initial_data = json.loads(state_str) if state_str and state_str != 'null' else {}
            rnr = initial_data.get("rnr", {}) if isinstance(initial_data, dict) else {}
            total_count = rnr.get("reviewsTopTotalCount", 0)
            for r in rnr.get("top_reviews", []):
                n = _bb_normalise_review(r)
                all_reviews[r["review_id"]] = n
            for r in rnr.get("reviews", []):
                n = _bb_normalise_review(r)
                all_reviews[r["review_id"]] = n
            print(f"[BB] From PRELOADED_STATE: {len(all_reviews)} reviews, total={total_count}")
        except Exception as e:
            print(f"[BB] PRELOADED_STATE error: {e}")

        # -- Paginate all review pages directly ---------------------------
        max_pages = max(150, (total_count // 10) + 5) if total_count else 150
        print(f"[BB] Will paginate up to {max_pages} pages...")

        def extract_reviews_from_state(state_str2):
            """Extract reviews from __PRELOADED_STATE__ string."""
            found = []
            try:
                pdata = json.loads(state_str2) if state_str2 and state_str2 != 'null' else {}
                prnr = pdata.get("rnr", {}) if isinstance(pdata, dict) else {}
                for r in prnr.get("top_reviews", []) + prnr.get("reviews", []):
                    if isinstance(r, dict) and "review_id" in r:
                        found.append(r)
            except Exception:
                pass
            return found

        async def get_state_reviews(pg_obj):
            """Try multiple times to get reviews from __PRELOADED_STATE__."""
            for attempt in range(3):
                s = await pg_obj.evaluate("""
                    () => {
                        try {
                            const s = window.__PRELOADED_STATE__;
                            if (!s) return null;
                            if (typeof s === 'string') return s;
                            return JSON.stringify(s);
                        } catch(e) { return null; }
                    }
                """)
                if s:
                    found = extract_reviews_from_state(s)
                    if found:
                        return found
                await asyncio.sleep(2)
            return []

        stale_pages = 0
        for pg in range(1, max_pages + 1):
            pg_url = f"{review_url_base}?page={pg}"
            try:
                await page.goto(pg_url, wait_until="domcontentloaded", timeout=60_000)
            except Exception as e:
                print(f"[BB] Page {pg} load error: {e}")
                stale_pages += 1
                if stale_pages >= 3:
                    break
                continue
            await asyncio.sleep(5)  # wait for JS hydration

            prev_count = len(all_reviews)

            # Primary: get reviews from __PRELOADED_STATE__
            state_reviews = await get_state_reviews(page)
            for r in state_reviews:
                all_reviews[r["review_id"]] = _bb_normalise_review(r)

            # Secondary: collect from network-intercepted API responses
            for entry in list(capture.api_calls_with_body):
                body = entry["body"]
                candidate_lists = []
                if isinstance(body, dict):
                    for key in ("reviews", "top_reviews", "review_list", "data"):
                        val = body.get(key)
                        if isinstance(val, list):
                            candidate_lists.extend(val)
                    rnr_block = body.get("rnr", {})
                    if isinstance(rnr_block, dict):
                        for key in ("reviews", "top_reviews"):
                            candidate_lists.extend(rnr_block.get(key, []))
                elif isinstance(body, list):
                    candidate_lists = body
                for r in candidate_lists:
                    if isinstance(r, dict) and "review_id" in r:
                        all_reviews[r["review_id"]] = _bb_normalise_review(r)

            added = len(all_reviews) - prev_count
            print(f"[BB] Page {pg}: +{added} reviews, total={len(all_reviews)}")

            if added == 0:
                stale_pages += 1
                if stale_pages >= 3:
                    print(f"[BB] 3 consecutive empty pages, stopping.")
                    break
            else:
                stale_pages = 0

            if _bb_reached_cutoff(all_reviews, until_date):
                print(f"[BB] Reached cutoff date, stopping.")
                break
            if total_count and len(all_reviews) >= total_count:
                print(f"[BB] Got all {total_count} reviews.")
                break

            await asyncio.sleep(1)  # polite delay between pages

        # (pagination handled above in page loop)

        # -- DOM fallback if nothing scraped yet ----------------------------
        if not all_reviews:
            for pg in range(1, 51):
                try:
                    await page.goto(
                        f"{review_url_base}?page={pg}",
                        wait_until="domcontentloaded", timeout=60_000
                    )
                    await asyncio.sleep(3)
                    dom_reviews = await page.evaluate("""
                    () => {
                        const results = [];
                        const selectors = [
                            '[class*="review"]', '[class*="Review"]',
                            '[class*="rating"]', '[data-review]',
                        ];
                        for (const sel of selectors) {
                            const els = document.querySelectorAll(sel);
                            for (const el of els) {
                                const txt = (el.innerText || '').trim();
                                if (txt.length > 30 && txt.length < 2000) {
                                    results.push({ review_description: txt, rating: 0 });
                                }
                            }
                            if (results.length > 3) break;
                        }
                        return results;
                    }
                    """)
                    new_count = 0
                    for r in (dom_reviews or []):
                        key = r.get("review_description", "")[:80]
                        if key and key not in all_reviews:
                            all_reviews[key] = {
                                "review_id": key,
                                "rating": r.get("rating", 0),
                                "review_header": "",
                                "review_text": r.get("review_description", ""),
                                "date": "",
                                "product_name": "",
                                "brand": "",
                                "sku": "",
                                "category": "",
                            }
                            new_count += 1
                    if new_count == 0:
                        break
                except Exception:
                    break

        await browser.close()

    # Filter by date and return
    all_reviews = _bb_filter_by_date(all_reviews, until_date)
    return summary, list(all_reviews.values()), total_count or len(all_reviews)


def build_bigbasket_rows(reviews: list, slug: str, product_url: str = "") -> list:
    rows = []
    for i, r in enumerate(reviews, 1):
        review_header = r.get("review_header") or ""
        review_text = r.get("review_text") or ""
        review_combine = f"{review_header} {review_text}".strip()
        review_id = make_review_id("BB", r.get("brand", "") or slug, slug, i)
        row = [
            review_id,
            r.get("rating", ""),
            review_header,
            review_text,
            review_combine,
            "india",
            r.get("date", ""),
            r.get("product_name", ""),
            r.get("brand", ""),
            "",                       # variant_me
            str(r.get("sku", "")),
            product_url,
            r.get("category", ""),
            "bigbasket",
        ]
        rows.append(row)
    return rows


def build_bigbasket_excel(reviews: list, summary: dict, slug: str, product_url: str = "") -> bytes:
    wb = Workbook()

    # Sheet 1: Reviews
    ws = wb.active
    ws.title = "BigBasket Reviews"
    build_sheet_header(ws, "1F4E79")
    rows_data = build_bigbasket_rows(reviews, slug, product_url)
    write_rows(ws, rows_data, rating_col_idx=2)
    apply_col_widths(ws)

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    summary_rows = [
        ("Product URL",    summary.get("product_url", product_url)),
        ("Overall Rating", summary.get("overall_rating", "")),
        ("Total Ratings",  summary.get("total_ratings", "")),
        ("Total Reviews",  summary.get("total_reviews", "")),
        ("5 Star",         summary.get("star_5", "")),
        ("4 Star",         summary.get("star_4", "")),
        ("3 Star",         summary.get("star_3", "")),
        ("2 Star",         summary.get("star_2", "")),
        ("1 Star",         summary.get("star_1", "")),
        ("Reviews Scraped", len(reviews)),
        ("Scrape Date",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for h in summary.get("highlights", []):
        summary_rows.append((f"{h['attribute']} Score",   h.get("score", "")))
        summary_rows.append((f"{h['attribute']} Ratings", h.get("ratings", "")))
    for r_idx, (field, val) in enumerate(summary_rows, 1):
        ws2.cell(row=r_idx, column=1, value=field)
        ws2.cell(row=r_idx, column=2, value=val)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ------------------------------------------------------------------------------
# FLIPKART
# ------------------------------------------------------------------------------

def parse_relative_date(raw: str):
    """Convert Flipkart relative dates ('5 days ago', '2 months ago') to datetime.
    From flipkart/scrap.py � handles relative + absolute formats.
    """
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    raw = raw.strip().lstrip("�").strip().lower()

    m = re.match(r"(\d+)\s+(day|week|month|year)s?\s+ago", raw)
    if m:
        n, unit = int(m.group(1)), m.group(2)
        if unit == "day":
            return today - timedelta(days=n)
        if unit == "week":
            return today - timedelta(weeks=n)
        if unit == "month":
            return today - timedelta(days=n * 30)
        if unit == "year":
            return today - timedelta(days=n * 365)

    for fmt in ("%d %b %Y", "%d %B %Y", "%b %d, %Y", "%b, %Y", "%b %Y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def parse_flipkart_url(url: str):
    m = re.search(r"flipkart\.com/([^/]+)/product-reviews/([^?/]+)", url)
    if m:
        return m.group(1), m.group(2), None, None
    m2 = re.search(r"flipkart\.com/([^/]+)/p/([^?/]+)", url)
    if m2:
        pid_m = re.search(r"pid=([A-Z0-9]+)", url)
        lid_m = re.search(r"lid=([A-Z0-9]+)", url)
        return m2.group(1), m2.group(2), pid_m.group(1) if pid_m else None, lid_m.group(1) if lid_m else None
    raise ValueError("Could not parse Flipkart URL. Please use a product or review page URL.")


# -- JS card extractor (enhanced from scraper-v2/main.py DUMP_JS) --------------

DUMP_JS = r"""
() => {
    const cards = [];
    const seen  = new Set();
    function getRating(cardEl) {
        for (const el of cardEl.querySelectorAll('[aria-label]')) {
            const lbl = (el.getAttribute('aria-label') || '').toLowerCase();
            const m = lbl.match(/(\d(?:\.\d)?)\s*(?:out of|\/)\s*5/)
                   || lbl.match(/rated\s*(\d(?:\.\d)?)/)
                   || lbl.match(/^(\d(?:\.\d)?)\s*star/);
            if (m) { const v = parseFloat(m[1]); if (v>=1&&v<=5) return Math.round(v); }
        }
        for (const el of cardEl.querySelectorAll('[data-score],[data-rating],[data-value]')) {
            const v = parseFloat(el.getAttribute('data-score')||el.getAttribute('data-rating')||el.getAttribute('data-value'));
            if (!isNaN(v) && v>=1 && v<=5) return Math.round(v);
        }
        for (const el of cardEl.querySelectorAll('span,div')) {
            if (el.children.length > 0) continue;
            const txt = (el.textContent || '').trim();
            if (!/^[1-5](\.[05])?$/.test(txt)) continue;
            const bg = window.getComputedStyle(el).backgroundColor || '';
            if (bg && bg !== 'rgba(0, 0, 0, 0)' && bg !== 'rgb(255, 255, 255)') {
                const v = parseFloat(txt); if (v>=1&&v<=5) return Math.round(v);
            }
        }
        for (const el of cardEl.querySelectorAll('span,div,b,strong')) {
            if (el.children.length > 0) continue;
            const txt = (el.textContent || '').trim();
            if (/^[1-5]$/.test(txt) && el.offsetWidth < 60 && el.offsetHeight < 60) return parseInt(txt);
        }
        return 0;
    }
    function getMediaLinks(cardEl) {
        const links = [];
        for (const img of cardEl.querySelectorAll('img')) {
            const src = img.getAttribute('src') || img.getAttribute('data-src') || '';
            if (src && src.startsWith('http') && !src.includes('profile') && !src.includes('avatar') && !src.includes('logo')) links.push(src);
        }
        for (const vid of cardEl.querySelectorAll('video source, video')) {
            const src = vid.getAttribute('src') || '';
            if (src && src.startsWith('http')) links.push(src);
        }
        for (const a of cardEl.querySelectorAll('a[href]')) {
            const href = a.getAttribute('href') || '';
            if (/\.(jpg|jpeg|png|gif|webp|mp4|mov)/i.test(href)) links.push(href);
        }
        return [...new Set(links)];
    }
    // Strategy 1: css-g5y9jx cards (current Flipkart layout)
    for (const card of document.querySelectorAll('div.css-g5y9jx')) {
        const s = card.getAttribute('style') || '';
        if (!s.includes('padding-left: 16px') || !s.includes('padding-top: 16px')) continue;
        const t = (card.innerText || '').trim();
        if (t.length < 40) continue;
        const key = t.slice(0, 150);
        if (!seen.has(key)) { seen.add(key); cards.push({ rating: getRating(card), text: t, media: getMediaLinks(card) }); }
    }
    if (cards.length >= 3) return JSON.stringify(cards);
    // Strategy 2: css-146c3p1 body elements
    cards.length = 0; seen.clear();
    for (const bodyEl of document.querySelectorAll('div.css-146c3p1')) {
        const bodyTxt = (bodyEl.innerText || '').trim();
        if (bodyTxt.length < 30) continue;
        let cur = bodyEl.parentElement;
        for (let i = 0; i < 10 && cur && cur !== document.body; i++) {
            const t = (cur.innerText || '').trim();
            if (t.length > 100 && t.length < 3000) {
                const key = t.slice(0, 150);
                if (!seen.has(key)) { seen.add(key); cards.push({ rating: getRating(cur), text: t, media: getMediaLinks(cur) }); }
                break;
            }
            cur = cur.parentElement;
        }
    }
    if (cards.length >= 3) return JSON.stringify(cards);
    // Strategy 3: any div containing a month name
    const MONTH_RE = /\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b/i;
    cards.length = 0; seen.clear();
    for (const el of document.querySelectorAll('div')) {
        const t = (el.innerText || '').trim();
        if (t.length < 80 || t.length > 3000 || !MONTH_RE.test(t)) continue;
        const childMatch = [...el.querySelectorAll('div')].some(d => {
            const dt = (d.innerText||'').trim();
            return dt.length >= 80 && dt.length <= 3000 && MONTH_RE.test(dt);
        });
        if (childMatch) continue;
        const key = t.slice(0, 150);
        if (!seen.has(key)) { seen.add(key); cards.push({ rating: getRating(el), text: t, media: getMediaLinks(el) }); }
    }
    return JSON.stringify(cards);
}
"""


JUNK_PATTERNS = re.compile(
    r"^(\.|{|/\*|rgba|transform|stroke|0h256|font-|letter-|line-|padding|margin|border|color:|background|align|display|flex|position|overflow|cursor|box-sizing)"
)


def parse_flipkart_card(rating, text, media=None):
    """Parse a single Flipkart review card text into structured fields.
    Logic from flipkart/scrap.py with improvements.
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    month_re = re.compile(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b", re.I)
    junk = {
        "READ MORE", "Certified Buyer", "Like", "Dislike",
        "Report Abuse", "HELPFUL", "Was this review helpful",
        "Thank you for your feedback"
    }

    def is_junk(l):
        return (l in junk or re.fullmatch(r"[\d,\.?\s]+", l) or len(l) < 2)

    # Find date line
    date_str = ""
    for l in lines:
        if month_re.search(l) and len(l) < 35:
            date_str = l
            break

    # Parse date ? YYYY-MM-DD
    parsed_date = ""
    if date_str:
        dt = parse_relative_date(date_str)
        if dt:
            parsed_date = dt.strftime("%Y-%m-%d")
        else:
            parsed_date = date_str.strip()

    # Extract rating from text if JS getRating returned 0
    if not rating or rating == 0:
        for l in lines:
            m_r = re.match(r"^([1-5])(\.\d)?$", l.strip())
            if m_r:
                rating = int(m_r.group(1))
                break

    # Find title � first non-junk, non-date line of reasonable length
    title = ""
    remaining_lines = [l for l in lines if not is_junk(l) and not month_re.search(l)]
    for l in remaining_lines:
        if 4 <= len(l) <= 140:
            title = l
            break

    # Body � everything after title, excluding junk and date lines
    body_lines, seen_b = [], set()
    for l in lines:
        if is_junk(l) or month_re.search(l) or l == title:
            continue
        # Skip "Review for:" variant lines and "Verified Purchase" etc
        if l.startswith("Review for:") or "Verified Purchase" in l:
            continue
        if len(l) > 8 and l not in seen_b:
            seen_b.add(l)
            body_lines.append(l)
    body = " ".join(body_lines)

    # Reviewer name � line just before date, short, not junk
    reviewer = ""
    if date_str:
        idx = next((i for i, l in enumerate(lines) if l == date_str), -1)
        for offset in [-1, -2]:
            if 0 <= idx + offset < len(lines):
                c = lines[idx + offset]
                # Reviewer names are short and don't contain digits/junk
                if 2 < len(c) < 40 and not is_junk(c) and not month_re.search(c) and not re.search(r"\d{4}", c):
                    reviewer = c
                    break

    helpful = 0
    m = re.search(r"(\d+)\s*(?:people found|helpful)", text, re.I)
    if m:
        helpful = int(m.group(1))

    return {
        "rating": rating,
        "title": title,
        "body": body,
        "reviewer": reviewer,
        "date_str": parsed_date,
        "certified": "certified buyer" in text.lower(),
        "helpful": helpful,
        "media": media or [],
    }


async def scrape_flipkart(url: str, until_date=None):
    """
    Scrape Flipkart reviews using DUMP_JS card extractor across paginated review pages.
    Sorts by Latest. Stops at until_date if provided.
    """
    slug, item, pid, lid = parse_flipkart_url(url)
    all_reviews: list[dict] = []
    seen_keys: set = set()

    def review_url(pg):
        base = f"https://www.flipkart.com/{slug}/product-reviews/{item}?marketplace=FLIPKART&sortOrder=MOST_RECENT&page={pg}"
        if pid:
            base += f"&pid={pid}"
        if lid:
            base += f"&lid={lid}"
        return base

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-IN",
            viewport={"width": 1280, "height": 900},
            extra_http_headers={"Accept-Language": "en-IN,en;q=0.9"},
        )
        await context.add_init_script("""
            Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
            window.chrome={runtime:{}};
            Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});
            Object.defineProperty(navigator,'languages',{get:()=>['en-IN','en']});
        """)
        page = await context.new_page()

        # Visit homepage, close any popup
        await page.goto("https://www.flipkart.com/", wait_until="domcontentloaded", timeout=60_000)
        await asyncio.sleep(random.uniform(3, 5))
        for txt in ["?", "�", "Close", "close"]:
            try:
                await page.get_by_text(txt, exact=True).first.click(timeout=1500)
                break
            except Exception:
                pass
        try:
            await page.keyboard.press("Escape")
        except Exception:
            pass

        consecutive_empty = 0
        for pg in range(1, 321):
            try:
                await page.goto(review_url(pg), wait_until="domcontentloaded", timeout=60_000)
            except Exception:
                consecutive_empty += 1
                if consecutive_empty >= 5:
                    break
                continue

            await asyncio.sleep(random.uniform(2.0, 3.5))

            # Scroll to trigger lazy loading
            for y in [0.25, 0.5, 0.75, 1.0]:
                await page.evaluate(f"window.scrollTo(0, document.body.scrollHeight*{y})")
                await asyncio.sleep(0.5)
            await asyncio.sleep(1.2)

            try:
                raw = await page.evaluate(DUMP_JS)
                cards = json.loads(raw) if raw else []
            except Exception:
                cards = []

            new = 0
            reached_cutoff = False
            for card in cards:
                r = parse_flipkart_card(card["rating"], card["text"], card.get("media", []))
                key = (r["reviewer"], r["body"][:200], r["date_str"], r["rating"])
                if key in seen_keys:
                    continue
                if not r["body"] and not r["title"] and not r["rating"]:
                    continue

                # Check until_date cutoff
                if until_date and r["date_str"]:
                    try:
                        review_d = datetime.strptime(r["date_str"], "%Y-%m-%d").date()
                        if review_d < until_date:
                            reached_cutoff = True
                            continue
                    except Exception:
                        pass

                seen_keys.add(key)
                new += 1
                all_reviews.append(r)

            if reached_cutoff:
                break

            if new == 0:
                consecutive_empty += 1
                if consecutive_empty >= 5:
                    break
            else:
                consecutive_empty = 0

        await browser.close()

    return all_reviews


def build_flipkart_rows(reviews: list, product_url: str = "") -> list:
    rows = []
    for i, r in enumerate(reviews, 1):
        review_header = r.get("title") or ""
        review_text = r.get("body") or ""
        review_combine = f"{review_header} {review_text}".strip()
        review_id = make_review_id("FK", "FL", "PR", i)
        row = [
            review_id,
            r["rating"],
            review_header,
            review_text,
            review_combine,
            "india",
            r["date_str"],
            "",   # product_name (not easily available from DOM)
            "",   # brand
            "",   # variant_me
            "",   # sku
            product_url,
            "",   # category
            "flipkart",
        ]
        rows.append(row)
    return rows


def build_flipkart_excel(reviews: list, product_url: str = "") -> bytes:
    wb = Workbook()

    # Sheet 1: Reviews
    ws = wb.active
    ws.title = "Flipkart Reviews"
    build_sheet_header(ws, "F47920")
    rows_data = build_flipkart_rows(reviews, product_url)
    write_rows(ws, rows_data, rating_col_idx=2)
    apply_col_widths(ws)

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    import pandas as pd
    df = [r["rating"] for r in reviews if r.get("rating")]
    avg = round(sum(df) / len(df), 2) if df else ""
    date_vals = sorted([r["date_str"] for r in reviews if r.get("date_str")])
    summary_rows = [
        ("Product URL",              product_url),
        ("Scrape Date",              datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Reviews Scraped",          len(reviews)),
        ("Average Rating (scraped)", avg),
        ("Date Range",               f"{date_vals[0]} ? {date_vals[-1]}" if date_vals else ""),
    ]
    for r_idx, (field, val) in enumerate(summary_rows, 1):
        ws2.cell(row=r_idx, column=1, value=field)
        ws2.cell(row=r_idx, column=2, value=val)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ------------------------------------------------------------------------------
# API Endpoints
# ------------------------------------------------------------------------------

@app.get("/health")
async def health():
    return {"status": "ok"}


@app.post("/api/validate-bigbasket")
async def validate_bigbasket(req: ScrapeRequest):
    try:
        sku, slug = parse_bigbasket_url(req.url)
        return {"sku": sku, "slug": slug, "valid": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/validate-flipkart")
async def validate_flipkart(req: ScrapeRequest):
    try:
        slug, item, pid, lid = parse_flipkart_url(req.url)
        return {"slug": slug, "item": item, "valid": True}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/scrape-bigbasket")
async def scrape_bigbasket_endpoint(req: ScrapeRequest):
    try:
        sku, slug = parse_bigbasket_url(req.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    until_date = parse_until_date(req.until_date) if req.until_date else None

    try:
        summary, reviews, total = await scrape_bigbasket(sku, slug, until_date=until_date)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scraping failed: {str(e)}")

    if not reviews:
        raise HTTPException(status_code=404, detail="No reviews found.")

    rows_data = build_bigbasket_rows(reviews, slug, req.url)
    save_reviews_to_supabase(rows_data)
    excel_bytes = build_bigbasket_excel(reviews, summary, slug, req.url)
    filename = f"BigBasket_{slug[:30]}_{len(reviews)}_reviews.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/scrape-flipkart")
async def scrape_flipkart_endpoint(req: ScrapeRequest):
    try:
        parse_flipkart_url(req.url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    until_date = parse_until_date(req.until_date) if req.until_date else None

    try:
        reviews = await scrape_flipkart(req.url, until_date=until_date)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scraping failed: {str(e)}")

    if not reviews:
        raise HTTPException(status_code=404, detail="No reviews found.")

    rows_data = build_flipkart_rows(reviews, req.url)
    save_reviews_to_supabase(rows_data)
    excel_bytes = build_flipkart_excel(reviews, req.url)
    filename = f"Flipkart_{len(reviews)}_reviews.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




